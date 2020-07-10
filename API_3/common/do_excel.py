import openpyxl
from API_3.common import httpRequest


class Case:
    def __init__(self):
        self.id = None
        self.title = None
        self.url = None
        self.method = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row #获取最大行数

        cases = [] #列表，用来存放所有的测试用例
        for r in range(2, max_row+1):
            # #方法一：用字典对单条用例的各个字段进行数据存储
            # case = {}
            # case["case_id"] = self.sheet.cell(row=r, column=1)
            # case["title"] = self.sheet.cell(row=r, column=2)
            case = Case()
            case.id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value

            cases.append(case)
        self.workbook.close()
        return cases

    def write_result(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()

# if __name__ == "__main__":
#     do_excel = DoExcel("../data/cases.xlsx", "login")
#     cases = do_excel.get_cases()
#     http_request = httpRequest.HttpRequest()
#     for case in cases:
#         #从excel中读取的data是字符串，而http请求时，data是字典，所以要将data转为字典类型
#         print(type(case.data))
#         resp = http_request.request(case.method, case.url, eval(case.data))
#         print(resp.status_code)
#         print(resp.text)
#         actual = resp.text
#
#         if case.expected == actual:
#             do_excel.write_result(case.id+1, actual, "PASS")
#         else:
#             do_excel.write_result(case.id+1, actual, "FAIL")
